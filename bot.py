import os
import sqlite3
import re
from telegram import Update
from telegram.ext import ApplicationBuilder, MessageHandler, CommandHandler, filters, ContextTypes
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# token vem do Railway
TOKEN = os.getenv("TOKEN")

# base de dados
conn = sqlite3.connect("stock.db", check_same_thread=False)
cursor = conn.cursor()

# criar tabelas
cursor.execute("""
CREATE TABLE IF NOT EXISTS produtos (
    codigo TEXT PRIMARY KEY,
    maquina TEXT,
    descricao TEXT,
    stock INTEGER DEFAULT 0,
    stock_confirmado INTEGER DEFAULT 0
)
""")

cursor.execute("""
CREATE TABLE IF NOT EXISTS material_pedir (
    codigo TEXT,
    maquina TEXT,
    descricao TEXT,
    stock INTEGER
)
""")

# exemplo inicial (podes mudar depois)
cursor.execute("""
INSERT OR IGNORE INTO produtos VALUES ('10','calandra 2','correia 100x160',5,5)
""")

conn.commit()

# verificar stock baixo
def verificar_stock_baixo(codigo):
    cursor.execute("SELECT codigo, maquina, descricao, stock FROM produtos WHERE codigo=?", (codigo,))
    produto = cursor.fetchone()

    if produto:
        codigo, maquina, descricao, stock = produto

        if stock < 3:
            cursor.execute("DELETE FROM material_pedir WHERE codigo=?", (codigo,))
            cursor.execute("INSERT INTO material_pedir VALUES (?,?,?,?)",
                           (codigo, maquina, descricao, stock))
        else:
            cursor.execute("DELETE FROM material_pedir WHERE codigo=?", (codigo,))

        conn.commit()

# tratar mensagens
async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()

    match_mov = re.match(r"(\d+)\s*([+-]\d+)", text)
    match_stock = re.match(r"(\d+)\s*(\d+)$", text)

    # movimento + ou -
    if match_mov:
        codigo = match_mov.group(1)
        movimento = int(match_mov.group(2))

        cursor.execute("SELECT maquina, descricao, stock, stock_confirmado FROM produtos WHERE codigo=?", (codigo,))
        r = cursor.fetchone()

        if not r:
            await update.message.reply_text("CÃ³digo nÃ£o encontrado")
            return

        maquina, descricao, stock, confirmado = r
        stock += movimento

        cursor.execute("UPDATE produtos SET stock=? WHERE codigo=?", (stock, codigo))
        conn.commit()

        verificar_stock_baixo(codigo)

        diferenca = confirmado - stock

        msg = f"""
ðŸ“¦ {codigo} ({maquina})
{descricao}

ðŸ”„ Movimento: {movimento}
ðŸ“Š Stock: {stock}
âœ… Confirmado: {confirmado}
ðŸ“‰ DiferenÃ§a: {diferenca}
"""

        if stock < 3:
            msg += "\nâš ï¸ STOCK BAIXO"

        if abs(diferenca) >= 3:
            msg += "\nðŸš¨ ERRO DE STOCK"

        await update.message.reply_text(msg)

    # stock confirmado
    elif match_stock:
        codigo = match_stock.group(1)
        confirmado = int(match_stock.group(2))

        cursor.execute("SELECT maquina, descricao, stock FROM produtos WHERE codigo=?", (codigo,))
        r = cursor.fetchone()

        if not r:
            await update.message.reply_text("CÃ³digo nÃ£o encontrado")
            return

        maquina, descricao, stock = r

        cursor.execute("UPDATE produtos SET stock_confirmado=? WHERE codigo=?", (confirmado, codigo))
        conn.commit()

        diferenca = confirmado - stock

        msg = f"""
âœ… STOCK CONFIRMADO

ðŸ“¦ {codigo} ({maquina})
{descricao}

ðŸ“Š Sistema: {stock}
âœ… Confirmado: {confirmado}
ðŸ“‰ DiferenÃ§a: {diferenca}
"""

        if abs(diferenca) >= 3:
            msg += "\nðŸš¨ ERRO GRAVE"

        await update.message.reply_text(msg)

    else:
        await update.message.reply_text(
            "Formato:\n"
            "10 +1\n"
            "10 -2\n"
            "10 5 (confirmado)"
        )

# exportar excel
async def export_excel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    df = pd.read_sql_query("SELECT * FROM produtos", conn)
    df["diferenca"] = df["stock_confirmado"] - df["stock"]

    file = "stock.xlsx"
    df.to_excel(file, index=False)

    wb = load_workbook(file)
    ws = wb.active

    red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    for row in range(2, ws.max_row + 1):
        if ws[f"D{row}"].value < 3:
            for col in range(1, 7):
                ws.cell(row=row, column=col).fill = red

    wb.save(file)

    await update.message.reply_document(open(file, "rb"))

# ver material a pedir
async def ver_pedidos(update: Update, context: ContextTypes.DEFAULT_TYPE):
    cursor.execute("SELECT * FROM material_pedir")
    rows = cursor.fetchall()

    if not rows:
        await update.message.reply_text("Sem material a pedir")
        return

    msg = "ðŸš¨ MATERIAL A PEDIR:\n\n"
    for r in rows:
        msg += f"{r[0]} | {r[1]} | stock: {r[3]}\n"

    await update.message.reply_text(msg)

# iniciar bot
app = ApplicationBuilder().token(TOKEN).build()

app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))
app.add_handler(CommandHandler("excel", export_excel))
app.add_handler(CommandHandler("pedidos", ver_pedidos))

app.run_polling()
